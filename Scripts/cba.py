from argparse import ArgumentParser
from collections import defaultdict
import numpy
import pandas
from pathlib import Path
from openpyxl import load_workbook
from tables.exceptions import NoSuchNodeError

import utils.config
import utils.log as log
import parameters.assignment as param
from datahandling.matrixdata import MatrixData


SCRIPT_DIR = Path(__file__).parent

VEHICLE_KMS_FILE = "vehicle_kms_vdfs.txt"
TRANSIT_KMS_FILE = "transit_kms.txt"
LINK_LENGTH_FILE = "link_lengths.txt"
NOISE_FILE = "noise_areas.txt"
STATION_FILE = "transit_stations.txt"

TRANSIT_AGGREGATIONS = {
    "bus": ("Bus", "Long_d_bus", "BRT"),
    "metro": ("Metro",),
    "train": ("Train", "Local_trai"),
    "tram": ("Tram", "Light_rail"),
}

TRANSLATIONS = {
    "car_work": "ha_tyo",
    "car_leisure": "ha_muu",
    "transit_work": "jl_tyo",
    "transit_leisure": "jl_muu",
    "airplane": "lento",
    "long_d_bus": "kaukob",
    "train": "juna",
    "bike": "pp_tyo",
    "bike_leisure": "pp_muu",
    "truck": "ka",
    "semi_trailer": "puolip",
    "trailer_truck": "yhd",
    "van": "pa",
}

CELL_INDICES = {
    "gains": {
        "cols": {
            "aht": 'B',
            "pt": 'C',
            "iht": 'D',
        },
        "rows": {
            # Different rows for the two years
            # One row for existing users, one row for additional users
            1: {
                "time": ("9", "10"),
                "dist": ("22", "23"),
                "cost": ("37", "38"),
            },
            2: {
                "time": ("14", "15"),
                "dist": ("27", "28"),
                "cost": ("42", "43"),
            },
        },
    },
    "transit_revenue": {
        "rows": {
            1: "43",
            2: "46",
        },
    },
    "car_revenue": {
        "cols": {
            "aht": 'F',
            "pt": 'G',
            "iht": 'H',
        },
        "rows": {
            1: "8",
            2: "13",
        },
    },
    "car_miles": {
        "cols": {
            0: {  # Before
                # Column index for each volume-delay function
                1: 'R',
                2: 'S',
                3: 'T',
                4: 'U',
                5: 'V',
            },
            1: {  # After
                1: "AA",
                2: "AB",
                3: "AC",
                4: "AD",
                5: "AE",
            },
        },
        "rows": {
            1: {
                "bus": "15",
                "car": "19",
                "van": "20",
                "truck": "21",
                "trailer_truck": "22",
            },
            2: {
                "bus": "28",
                "car": "32",
                "van": "33",
                "truck": "34",
                "trailer_truck": "35",
            },
        },
    },
    "transit_miles": {
        "cols": {
            "dist": 'S',
            "time": 'T',
        },
        "rows": {
            1: {
                "bus": "8",
                "tram": "10",
                "metro": "11",
                "train": "12",
            },
            2: {
                "bus": "16",
                "tram": "18",
                "metro": "19",
                "train": "20",
            },
        },
    },
    "noise": {
        1: "I24",
        2: "I36",
    },
    "transit_stations": {
        1: {
            "metro": "U11",
            "train": "U12",
        },
        2: {
            "metro": "U19",
            "train": "U20",
        },
    },
    "link_lengths": {
        1: {
            "motorway": "G73",
            "multi-lane": "G74",
            "single-lane": "G75",
            "train": "G76",
            "metro": "G77",
            "tram": "G78",
        },
        2: {
            "motorway": "G82",
            "multi-lane": "G83",
            "single-lane": "G84",
            "train": "G85",
            "metro": "G86",
            "tram": "G87",
        },
    },
}

def run_cost_benefit_analysis(scenario_0, scenario_1, year, workbook, submodel):
    """Runs CBA and writes the results to excel file.

    Parameters
    ----------
    scenario_0 : str
        Name of do-nothing scenario, for which
        forecast results are available in Results folder
    scenario_1 : str
        Name of project scenario, for which
        forecast results are available in Results folder
    year : int
        The evaluation year (1 or 2)
    results_directory : str
        Path to where "scenario_name/Matrices" result folder exists
    workbook : openpyxl.WorkBook
        The excel workbook where to save results

    Returns
    -------
    pandas.DataFrame
        Table of zone-wise consumer surplus results
        (travel time, travel cost, distance, revenue)
    """
    if year not in (1, 2):
        raise ValueError("Evaluation year must be either 1 or 2")
    log.info("Analyse year {}...".format(year))

    # Calculate mile differences
    miles = []
    for scen in (scenario_0, scenario_1):
        df = read(VEHICLE_KMS_FILE, scen).set_index(["class", "v/d-func"])
        miles.append(df["veh_km"].unstack(level=0))
    for mile in miles:
        mile["car"] = mile["car_work"] + mile["car_leisure"]
    ws = workbook["Ulkoisvaikutukset"]
    cols = CELL_INDICES["car_miles"]["cols"]
    rows = CELL_INDICES["car_miles"]["rows"][year]
    for mode in rows:
        for vdf in cols[0]:
            ws[cols[0][vdf]+rows[mode]] = miles[0][mode][vdf]
            ws[cols[1][vdf]+rows[mode]] = miles[1][mode][vdf]

    # Calculate noise effect difference
    noises = []
    for scen in (scenario_0, scenario_1):
        noises.append(read(NOISE_FILE, scen).set_index("area"))
    noise_diff = noises[1] - noises[0]
    ws[CELL_INDICES["noise"][year]] = sum(noise_diff["population"])

    # Calculate transit mile differences
    transit_mile_diff = (read(TRANSIT_KMS_FILE, scenario_1)
                         - read(TRANSIT_KMS_FILE, scenario_0))
    for mode in TRANSIT_AGGREGATIONS:
        transit_mile_diff.loc[mode] = 0
        for submode in TRANSIT_AGGREGATIONS[mode]:
            transit_mile_diff.loc[mode] += transit_mile_diff.loc[submode]
    ws = workbook["Tuottajahyodyt"]
    cols = CELL_INDICES["transit_miles"]["cols"]
    rows = CELL_INDICES["transit_miles"]["rows"][year]
    for mode in rows:
        for imp_type in cols:
            ws[cols[imp_type]+rows[mode]] = transit_mile_diff[imp_type][mode]
    log.info("Mileage differences calculated")

    # Calculate link length differences
    linklength_diff = (read(LINK_LENGTH_FILE, scenario_1)
                       - read(LINK_LENGTH_FILE, scenario_0))
    indices = CELL_INDICES["link_lengths"][year]
    for linktype in indices:
        ws[indices[linktype]] = linklength_diff["length"][linktype]

    # Calculate transit station differences
    station_diff = (read(STATION_FILE, scenario_1)
                    - read(STATION_FILE, scenario_0))
    indices = CELL_INDICES["transit_stations"][year]
    for mode in indices:
        ws[indices[mode]] = station_diff["number"][mode]

    # Calculate gains and revenues
    results = defaultdict(float)
    for timeperiod in ["aht", "pt", "iht"]:
        data = {
            "scen_1": MatrixData(Path(scenario_1, "Matrices", submodel)),
            "scen_0": MatrixData(Path(scenario_0, "Matrices", submodel)),
        }
        revenues_transit = 0
        revenues_car = 0
        cols = CELL_INDICES["gains"]["cols"]
        rows = CELL_INDICES["gains"]["rows"][year]
        for transport_class in param.transport_classes:
            vol_fac = param.volume_factors[transport_class][timeperiod]
            demand = {}
            for scenario in data:
                with data[scenario].open("demand", timeperiod) as mtx:
                    demand[scenario] = mtx[transport_class]
                    zone_numbers = mtx.zone_numbers
                result_type = f"{transport_class}_demand_{scenario}"
                results[result_type] += vol_fac * demand[scenario].sum(0)
            for mtx_type in ["dist", "time", "cost", "toll_cost"]:
                cost = {}
                for scenario in data:
                    with data[scenario].open(mtx_type, timeperiod) as mtx:
                        try:
                            cost[scenario] = mtx[transport_class]
                            matrices_found = True
                        except NoSuchNodeError:
                            matrices_found = False
                if not matrices_found:
                    continue
                if transport_class == "train" and mtx_type == "dist":
                    # Max travel time with fixed 20 kmph speed
                    # and one-hour start time
                    maxtime = {scen: 3*cost[scen] + 60 for scen in data}
                if transport_class == "train" and mtx_type == "time":
                    for scen in cost:
                        cost[scen] = numpy.minimum(cost[scen], maxtime[scen])
                gains_existing, gains_additional = calc_gains(demand, cost)
                result_type = transport_class + "_" + mtx_type
                results[result_type] += (vol_fac *
                                         (gains_existing+gains_additional))
                if (mtx_type == "cost"
                        and transport_class in param.transit_classes):
                    revenue = calc_revenue(demand, cost)
                    revenues_transit += revenue
                    results["transit_revenue"] += vol_fac * revenue
                if mtx_type == "toll_cost":
                    revenue = calc_revenue(demand, cost)
                    revenues_car += revenue
                    results["car_revenue"] += vol_fac * revenue
                else:
                    ws = workbook[TRANSLATIONS[transport_class]]
                    ws[cols[timeperiod]+rows[mtx_type][0]] = gains_existing.sum()
                    ws[cols[timeperiod]+rows[mtx_type][1]] = gains_additional.sum()
            log.info(f"Mode {transport_class} calculated for {timeperiod}")
        ws = workbook["Tuottajahyodyt"]
        rows = CELL_INDICES["transit_revenue"]["rows"][year]
        ws[cols[timeperiod]+rows] = revenues_transit.sum()
        ws = workbook["Julkistaloudelliset"]
        cols = CELL_INDICES["car_revenue"]["cols"]
        rows = CELL_INDICES["car_revenue"]["rows"][year]
        try:
            ws[cols[timeperiod]+rows] = revenues_car.sum()
        except AttributeError:
            pass
        log.info("Gains and revenues calculated for {}".format(timeperiod))
    log.info("Year {} completed".format(year))
    return pandas.DataFrame(results, zone_numbers)


def read(file_name, scenario_path):
    """Read data from file."""
    return pandas.read_csv(
        Path(scenario_path, file_name), delim_whitespace=True)


def calc_gains(demands, costs):
    """Calculate difference in consumer surplus between scen_1 and scen_0.

    Parameters
    ----------
    demands : dict
        scen_0 : numpy.ndarray
            Demand matrix for scenario 0
        scen_1 : numpy.ndarray
            Demand matrix for scenario 1
    costs : dict
        scen_0 : numpy.ndarray
            Impedance matrix for scenario 0
        scen_1 : numpy.ndarray
            Impedance matrix for scenario 1

    Returns
    -------
    numpy.ndarray
        Calculated gain for existing users per zone
    numpy.ndarray
        Calculated gain for new or evicted users per zone
    """
    gain = costs["scen_1"] - costs["scen_0"]
    demand_change = demands["scen_1"] - demands["scen_0"]
    demand_incr = demand_change >= 0
    demand_decr = demand_change < 0
    gains_existing = numpy.zeros_like(demand_change)
    gains_existing[demand_incr] = (demands["scen_0"]*gain)[demand_incr]
    gains_existing[demand_decr] = (demands["scen_1"]*gain)[demand_decr]
    gains_additional = numpy.zeros_like(demand_change)
    gains_additional[demand_incr] = 0.5*(demand_change*gain)[demand_incr]
    gains_additional[demand_decr] = -0.5*(demand_change*gain)[demand_decr]
    return gains_existing.sum(0), gains_additional.sum(0)


def calc_revenue(demands, costs):
    """Calculate difference in producer revenue between scen_1 and scen_0.

    Parameters
    ----------
    demands : dict
        scen_0 : numpy.ndarray
            Demand matrix for scenario 0
        scen_1 : numpy.ndarray
            Demand matrix for scenario 1
    costs : dict
        scen_0 : numpy.ndarray
            Impedance matrix for scenario 0
        scen_1 : numpy.ndarray
            Impedance matrix for scenario 1

    Returns
    -------
    numpy.ndarray
        Calculated revenue per zone
    """
    demand_change = demands["scen_1"] - demands["scen_0"]
    demand_incr = demand_change >= 0
    demand_decr = demand_change < 0
    cost_change = costs["scen_1"] - costs["scen_0"]
    revenue = numpy.zeros_like(demand_change)
    revenue[demand_incr] = ((costs["scen_1"]*demand_change)[demand_incr]
                            + (cost_change*demands["scen_0"])[demand_incr])
    revenue[demand_decr] = ((costs["scen_0"]*demand_change)[demand_decr]
                            + (cost_change*demands["scen_1"])[demand_decr])
    return revenue.sum(0)


if __name__ == "__main__":
    config = utils.config.read_from_file()
    parser = ArgumentParser(epilog="Calculates the Cost-Benefit Analysis between Results of two HELMET-Scenarios, "
                                   "and writes the outcome in CBA_kehikko.xlsx -file (in same folder).")
    parser.add_argument(
        "baseline_scenario", type=str, help="A 'do-nothing' baseline scenario")
    parser.add_argument(
        "projected_scenario", type=str,
        help="A projected scenario, compared to the baseline scenario")
    parser.add_argument(
        "baseline_scenario_2", nargs='?', type=str,
        help="A 'do-nothing' baseline scenario for second forecast year (optional)")
    parser.add_argument(
        "projected_scenario_2", nargs='?', type=str,
        help="A projected scenario, compared to the baseline scenario for second forecast year (optional)")
    parser.add_argument(
        "--log-format",
        choices={"TEXT", "JSON"},
        default="JSON",
    )
    parser.add_argument(
        "--log-level",
        choices={"DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"},
        default=config["LOG_LEVEL"],
    )
    parser.add_argument(
        "--scenario-name",
        type=str,
        default=config["SCENARIO_NAME"],
        help="Name of HELMET scenario. Influences result folder name and log file name."),
    parser.add_argument(
        "--results-path", type=str, required=True,
        help="Path to Results directory.")
    parser.add_argument(
        "--submodel",
        type=str,
        default=config["SUBMODEL"],
        help="Name of submodel, used for choosing appropriate zone mapping"),
    args = parser.parse_args()
    log.initialize(args)
    wb = load_workbook(SCRIPT_DIR / "CBA_kehikko.xlsx")
    results = run_cost_benefit_analysis(
        args.baseline_scenario, args.projected_scenario, 1, wb, args.submodel)
    if (args.baseline_scenario_2 is not None
            and args.baseline_scenario_2 != "undefined"):
        run_cost_benefit_analysis(
            args.baseline_scenario_2, args.projected_scenario_2, 2, wb, args.submodel)
    results_filename = "cba_{}_{}".format(
        Path(args.projected_scenario).name,
        Path(args.baseline_scenario).name)
    wb.save(Path(args.results_path, results_filename + ".xlsx"))
    results.to_csv(
        Path(args.results_path, results_filename + ".txt"),
        sep='\t', float_format="%8.1f")
    log.info("CBA results saved to file: {}".format(results_filename))
